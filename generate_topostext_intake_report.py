#!/usr/bin/env python3
"""
Generate a reviewable intake report for Brady's ToposText Stephanus HTML.

This is intentionally read-only with respect to the database. It parses the
latest fetched snapshot, classifies inline entity IDs using the same broad
conventions as the existing Brady import tooling, and writes an HTML report
plus a complete mention CSV for external review.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html as html_lib
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, replace
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import quote, unquote

from bs4 import BeautifulSoup, Tag


DEFAULT_SOURCE_NAME = "topostext_stephanus_html"
DEFAULT_OUTPUT = Path("exports/topostext_intake_report.html")
DEFAULT_PAULY_WORKBOOK_NAME = "PaulyHeadwordstoWikidata from Margherita scrape.xlsx"
KNOWN_ENTITY_TAGS = {"prn", "place", "person", "ethnic", "demonym", "patnym"}
SOURCE_TAG_FIXES = {
    "ppn": "prn",
    "demonymn": "demonym",
}

QID_RE = re.compile(r"^Q\d+$", re.IGNORECASE)
PLEIADES_RE = re.compile(r"^p?(\d+)$", re.IGNORECASE)
RE_RE = re.compile(r"^RE:.+$", re.IGNORECASE)
TOPOS_RE = re.compile(r"^\d+[A-Za-z].+$")
LOCAL_BRADY_RE = re.compile(r"^JBK\d+$", re.IGNORECASE)
PARAGRAPH_START_RE = re.compile(r"<p\b(?P<attrs>[^>]*)>", re.IGNORECASE | re.DOTALL)
ATTR_RE = re.compile(r"([A-Za-z_:][-A-Za-z0-9_:.]*)\s*=\s*(['\"])(.*?)\2", re.DOTALL)
OPEN_TAG_RE = re.compile(r"<(?P<tag>[A-Za-z][A-Za-z0-9]*)\b(?P<attrs>[^>]*)>", re.IGNORECASE | re.DOTALL)
NULL_CELL_VALUES = {"", "#n/a", "n/a", "none", "null", "nan"}


AUTHORITY_CLASS_LABELS = {
    "wikidata": "Wikidata QID",
    "pleiades_numeric": "numeric/Pleiades-like",
    "topostext_like": "ToposText-like",
    "re": "RE namespace",
    "yy_placeholder": "YY placeholder",
    "jj_placeholder": "JJ placeholder",
    "zzz": "zzz unresolved",
    "brady_local": "Brady/local",
    "missing": "missing id",
    "other": "other",
}


AUTHORITY_CLASS_NOTES = {
    "wikidata": "Q-prefixed Wikidata authority IDs.",
    "pleiades_numeric": "Bare numeric IDs; existing Brady import code treats these as Pleiades-style candidates.",
    "topostext_like": "Mixed numeric/alphabetic IDs already linked elsewhere in this repo as ToposText place IDs.",
    "re": "RE-prefixed identifiers; Brady confirmed these should be their own authority namespace.",
    "yy_placeholder": "Brady did a quick pattern search for a ToposText ID, but did not search Wikidata, RE, or Pleiades by hand.",
    "jj_placeholder": "Brady searched and found nothing; create a fresh ToposText-style ID from the best guess of what and where it is.",
    "zzz": "Unresolved inline placeholder, mostly on PRN mentions in the current snapshot.",
    "brady_local": "Local/person-specific identifier seen in the introductory matter.",
    "missing": "Tag has no usable id attribute.",
    "other": "Non-empty ID that does not match the current guessed patterns.",
}


@dataclass(frozen=True)
class ReEnrichment:
    re_id: str
    short_definition: str = ""
    article_item: str = ""
    subject_item: str = ""
    subject_label: str = ""
    wikipedia: str = ""
    title_redirect: str = ""
    volume: str = ""
    page: str = ""
    author: str = ""
    match_source: str = ""


@dataclass(frozen=True)
class SnapshotMetadata:
    snapshot_id: int | None
    source_name: str
    source_kind: str
    status: str
    local_path: str
    expected_name: str
    byte_count: int
    sha256: str
    fetched_at: str
    unchanged_from_snapshot_id: int | None = None


@dataclass(frozen=True)
class Entry:
    sequence: int
    work: str
    paragraph_id: str
    entry_key: str
    title: str
    wdate: str
    edate: str
    text: str


@dataclass(frozen=True)
class Mention:
    sequence: int
    entry_sequence: int
    work: str
    paragraph_id: str
    entry_key: str
    entry_title: str
    wdate: str
    edate: str
    tag_name: str
    original_tag_name: str
    tag_id: str
    authority_class: str
    mention_text: str
    authority_url: str
    context: str
    re_namespace_id: str = ""
    re_short_definition: str = ""
    re_article_item: str = ""
    re_subject_item: str = ""
    re_subject_label: str = ""
    re_author: str = ""
    re_volume: str = ""
    re_page: str = ""
    re_match_source: str = ""


@dataclass(frozen=True)
class ParsedToposText:
    entries: list[Entry]
    mentions: list[Mention]


def normalize_space(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def clean_cell(value: object) -> str:
    text = normalize_space(value)
    return "" if text.casefold() in NULL_CELL_VALUES else text


def canonical_tag_name(raw_tag_name: str) -> str:
    tag_name = normalize_space(raw_tag_name).casefold()
    return SOURCE_TAG_FIXES.get(tag_name, tag_name)


def classify_authority_id(raw_value: str) -> str:
    value = clean_cell(raw_value)
    if not value:
        return "missing"
    upper_value = value.upper()
    if upper_value == "ZZZ":
        return "zzz"
    if upper_value.endswith("YY"):
        return "yy_placeholder"
    if upper_value.endswith("JJ"):
        return "jj_placeholder"
    if RE_RE.fullmatch(value):
        return "re"
    if QID_RE.fullmatch(value):
        return "wikidata"
    if PLEIADES_RE.fullmatch(value):
        return "pleiades_numeric"
    if TOPOS_RE.fullmatch(value):
        return "topostext_like"
    if LOCAL_BRADY_RE.fullmatch(value):
        return "brady_local"
    return "other"


def authority_url(authority_class: str, raw_value: str) -> str:
    value = normalize_space(raw_value)
    if not value:
        return ""
    if authority_class == "wikidata":
        return f"https://www.wikidata.org/wiki/{quote(value.upper(), safe='')}"
    if authority_class == "pleiades_numeric":
        match = PLEIADES_RE.fullmatch(value)
        numeric_id = match.group(1) if match else value
        return f"https://pleiades.stoa.org/places/{quote(numeric_id, safe='')}"
    if authority_class == "topostext_like":
        return f"https://topostext.org/place/{quote(value, safe='')}"
    if authority_class == "re":
        return f"https://de.wikisource.org/wiki/{quote(value, safe=':/#(),%')}"
    return ""


def normalize_wikidata_url(raw_value: object) -> str:
    value = normalize_space(raw_value)
    if not value:
        return ""
    match = re.search(r"\bQ\d+\b", value, flags=re.IGNORECASE)
    if not match:
        return ""
    return f"https://www.wikidata.org/wiki/{match.group(0).upper()}"


def normalize_re_id(raw_value: str) -> str:
    value = normalize_space(raw_value)
    if not value:
        return ""
    if value.casefold().startswith("re:"):
        body = value[3:]
    else:
        body = value
    body = unquote(body).strip()
    body = re.sub(r"\s+", "_", body)
    return f"RE:{body}" if body else ""


def re_lookup_keys(raw_value: str) -> list[str]:
    normalized = normalize_re_id(raw_value)
    if not normalized:
        return []
    body = normalized[3:]
    bodies = {body}
    if "#" in body:
        bodies.add(body.split("#", 1)[0])
    expanded = set(bodies)
    for candidate in list(bodies):
        if "#" in candidate:
            continue
        match = re.fullmatch(r"(.+?)(\d+[a-z]?)", candidate, flags=re.IGNORECASE)
        if match and not match.group(1).endswith("_"):
            expanded.add(f"{match.group(1)}_{match.group(2)}")
        expanded.add(candidate.replace("_", " "))
    keys = [f"RE:{candidate}" for candidate in expanded if candidate]
    return sorted(set(keys), key=lambda item: (item != normalized, item))


def derive_entry_title(paragraph: Tag, full_text: str) -> str:
    direct_text = normalize_space(" ".join(paragraph.find_all(string=True, recursive=False)))
    if ":" in direct_text:
        candidate = normalize_space(direct_text.split(":", 1)[0])
        if candidate:
            return candidate[:120]
    if ":" in full_text:
        candidate = normalize_space(full_text.split(":", 1)[0])
        if candidate:
            return candidate[:120]
    return full_text[:120]


def context_snippet(full_text: str, focus_text: str, max_chars: int = 260) -> str:
    full_text = normalize_space(full_text)
    focus_text = normalize_space(focus_text)
    if not full_text:
        return ""
    start = 0
    if focus_text:
        index = full_text.find(focus_text)
        if index >= 0:
            start = max(0, index - 90)
    end = min(len(full_text), start + max_chars)
    if end - start < max_chars and start > 0:
        start = max(0, end - max_chars)
    snippet = full_text[start:end]
    if start > 0:
        snippet = "..." + snippet
    if end < len(full_text):
        snippet += "..."
    return snippet


def parse_attrs(raw_attrs: str) -> dict[str, str]:
    return {match.group(1).casefold(): match.group(3) for match in ATTR_RE.finditer(raw_attrs or "")}


def iter_raw_paragraphs(content: str) -> Iterable[tuple[dict[str, str], str]]:
    starts = list(PARAGRAPH_START_RE.finditer(content))
    for index, match in enumerate(starts):
        end = starts[index + 1].start() if index + 1 < len(starts) else len(content)
        raw_paragraph = content[match.start() : end]
        last_close = raw_paragraph.casefold().rfind("</p>")
        if last_close >= 0:
            raw_paragraph = raw_paragraph[: last_close + len("</p>")]
        yield parse_attrs(match.group("attrs")), raw_paragraph


def strip_markup(value: str) -> str:
    return normalize_space(BeautifulSoup(value or "", "html.parser").get_text(" ", strip=True))


def raw_mention_parts(raw_paragraph: str) -> Iterable[tuple[str, str, str, str]]:
    for match in OPEN_TAG_RE.finditer(raw_paragraph):
        original_tag_name = normalize_space(match.group("tag")).casefold()
        tag_name = canonical_tag_name(original_tag_name)
        if tag_name == "p":
            continue
        attrs = parse_attrs(match.group("attrs"))
        if tag_name not in KNOWN_ENTITY_TAGS and "id" not in attrs:
            continue

        close_re = re.compile(rf"</\s*{re.escape(original_tag_name)}\s*>", re.IGNORECASE)
        close_match = close_re.search(raw_paragraph, match.end())
        inner_markup = raw_paragraph[match.end() : close_match.start()] if close_match else ""
        yield tag_name, original_tag_name, normalize_space(attrs.get("id")), strip_markup(inner_markup)


def parse_topostext_html(content: str) -> ParsedToposText:
    raw_paragraphs = list(iter_raw_paragraphs(content))
    if not raw_paragraphs:
        soup = BeautifulSoup(content, "html.parser")
        raw_paragraphs = [
            ({name.casefold(): str(value) for name, value in paragraph.attrs.items()}, str(paragraph))
            for paragraph in soup.find_all("p")
            if isinstance(paragraph, Tag)
        ]

    entries: list[Entry] = []
    mentions: list[Mention] = []

    for entry_sequence, (raw_attrs, raw_paragraph) in enumerate(raw_paragraphs, start=1):
        fragment = BeautifulSoup(raw_paragraph, "html.parser")
        paragraph = fragment.find("p")
        if not isinstance(paragraph, Tag):
            continue
        work = normalize_space(raw_attrs.get("work") or paragraph.get("work"))
        paragraph_id = normalize_space(raw_attrs.get("id") or paragraph.get("id"))
        wdate = normalize_space(raw_attrs.get("wdate") or paragraph.get("wdate"))
        edate = normalize_space(raw_attrs.get("edate") or paragraph.get("edate"))
        entry_key = f"{work}:{paragraph_id}" if work or paragraph_id else f"paragraph:{entry_sequence}"
        full_text = normalize_space(paragraph.get_text(" ", strip=True))
        title = derive_entry_title(paragraph, full_text)
        entry = Entry(
            sequence=entry_sequence,
            work=work,
            paragraph_id=paragraph_id,
            entry_key=entry_key,
            title=title,
            wdate=wdate,
            edate=edate,
            text=full_text,
        )
        entries.append(entry)

        for tag_name, original_tag_name, tag_id, mention_text in raw_mention_parts(raw_paragraph):
            authority_class = classify_authority_id(tag_id)
            mentions.append(
                Mention(
                    sequence=len(mentions) + 1,
                    entry_sequence=entry_sequence,
                    work=work,
                    paragraph_id=paragraph_id,
                    entry_key=entry_key,
                    entry_title=title,
                    wdate=wdate,
                    edate=edate,
                    tag_name=tag_name,
                    original_tag_name=original_tag_name,
                    tag_id=tag_id,
                    authority_class=authority_class,
                    mention_text=mention_text,
                    authority_url=authority_url(authority_class, tag_id),
                    context=context_snippet(full_text, mention_text),
                )
            )

    return ParsedToposText(entries=entries, mentions=mentions)


def latest_snapshot_from_db(source_name: str, snapshot_id: int | None = None) -> SnapshotMetadata:
    from db import get_connection

    where_clause = "id = %s" if snapshot_id is not None else "source_name = %s AND status IN ('fetched', 'unchanged')"
    param = snapshot_id if snapshot_id is not None else source_name
    order_clause = "" if snapshot_id is not None else "ORDER BY fetched_at DESC, id DESC LIMIT 1"

    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT
                id,
                source_name,
                source_kind,
                status,
                COALESCE(local_path, '') AS local_path,
                COALESCE(expected_name, '') AS expected_name,
                byte_count,
                COALESCE(sha256, '') AS sha256,
                fetched_at,
                unchanged_from_snapshot_id
            FROM entity_source_snapshots
            WHERE {where_clause}
            {order_clause}
            """,
            (param,),
        )
        row = cur.fetchone()
    finally:
        conn.close()

    if not row:
        if snapshot_id is not None:
            raise RuntimeError(f"No entity_source_snapshots row found for id={snapshot_id}")
        raise RuntimeError(f"No fetched ToposText snapshot found for source_name={source_name!r}")

    return SnapshotMetadata(
        snapshot_id=int(row[0]),
        source_name=str(row[1]),
        source_kind=str(row[2]),
        status=str(row[3]),
        local_path=str(row[4]),
        expected_name=str(row[5]),
        byte_count=int(row[6] or 0),
        sha256=str(row[7]),
        fetched_at=row[8].isoformat() if hasattr(row[8], "isoformat") else str(row[8]),
        unchanged_from_snapshot_id=int(row[9]) if row[9] is not None else None,
    )


def metadata_from_file(path: Path, source_name: str) -> SnapshotMetadata:
    content = path.read_bytes()
    return SnapshotMetadata(
        snapshot_id=None,
        source_name=source_name,
        source_kind="local_file",
        status="local_file",
        local_path=str(path),
        expected_name=path.name,
        byte_count=len(content),
        sha256=hashlib.sha256(content).hexdigest(),
        fetched_at="",
    )


def resolve_snapshot_path(metadata: SnapshotMetadata) -> Path:
    if not metadata.local_path:
        raise RuntimeError("Snapshot row has no local_path")
    path = Path(metadata.local_path).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    if not path.exists():
        raise RuntimeError(f"Snapshot file does not exist: {path}")
    return path


def count_by_tag_and_class(mentions: Iterable[Mention]) -> dict[str, Counter[str]]:
    counts: dict[str, Counter[str]] = defaultdict(Counter)
    for mention in mentions:
        counts[mention.tag_name][mention.authority_class] += 1
    return counts


def default_pauly_workbook_candidates() -> list[Path]:
    return [
        Path("data/pauly") / DEFAULT_PAULY_WORKBOOK_NAME,
        Path.home() / "Downloads" / DEFAULT_PAULY_WORKBOOK_NAME,
    ]


def find_default_pauly_workbook() -> Path | None:
    for candidate in default_pauly_workbook_candidates():
        candidate = candidate.expanduser()
        if candidate.exists():
            return candidate
    return None


def update_re_enrichment(
    table: dict[str, ReEnrichment],
    raw_id: object,
    *,
    short_definition: object = "",
    article_item: object = "",
    subject_item: object = "",
    subject_label: object = "",
    wikipedia: object = "",
    title_redirect: object = "",
    volume: object = "",
    page: object = "",
    author: object = "",
    match_source: str,
) -> None:
    keys = re_lookup_keys(normalize_space(raw_id))
    if not keys:
        return
    re_id = keys[0]
    current = table.get(re_id, ReEnrichment(re_id=re_id))
    incoming = ReEnrichment(
        re_id=current.re_id or re_id,
        short_definition=clean_cell(short_definition) or current.short_definition,
        article_item=normalize_wikidata_url(article_item) or current.article_item,
        subject_item=normalize_wikidata_url(subject_item) or current.subject_item,
        subject_label=clean_cell(subject_label) or current.subject_label,
        wikipedia=clean_cell(wikipedia) or current.wikipedia,
        title_redirect=clean_cell(title_redirect) or current.title_redirect,
        volume=clean_cell(volume) or current.volume,
        page=clean_cell(page) or current.page,
        author=clean_cell(author) or current.author,
        match_source=current.match_source or match_source,
    )
    for key in keys:
        existing = table.get(key)
        if existing is None:
            table[key] = incoming
        else:
            table[key] = ReEnrichment(
                re_id=existing.re_id or incoming.re_id,
                short_definition=existing.short_definition or incoming.short_definition,
                article_item=existing.article_item or incoming.article_item,
                subject_item=existing.subject_item or incoming.subject_item,
                subject_label=existing.subject_label or incoming.subject_label,
                wikipedia=existing.wikipedia or incoming.wikipedia,
                title_redirect=existing.title_redirect or incoming.title_redirect,
                volume=existing.volume or incoming.volume,
                page=existing.page or incoming.page,
                author=existing.author or incoming.author,
                match_source=existing.match_source or incoming.match_source,
            )


def row_dict(headers: list[str], row: tuple[object, ...]) -> dict[str, object]:
    return {headers[index]: value for index, value in enumerate(row) if index < len(headers)}


def load_pauly_re_enrichment(path: Path) -> dict[str, ReEnrichment]:
    from openpyxl import load_workbook

    workbook = load_workbook(path.expanduser(), read_only=True, data_only=True)
    table: dict[str, ReEnrichment] = {}

    if "scrape_re_wikipedia_wikidata_li" in workbook.sheetnames:
        sheet = workbook["scrape_re_wikipedia_wikidata_li"]
        headers = [normalize_space(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = row_dict(headers, row)
            subject_item = data.get("wikidata subject item") or data.get("StatementSubject of")
            update_re_enrichment(
                table,
                data.get("RE_Title"),
                short_definition=data.get("ShortDefinition"),
                article_item=data.get("RE_ArticleItem"),
                subject_item=subject_item,
                wikipedia=data.get("wikipedia"),
                title_redirect=data.get("title_redirect"),
                match_source="scrape_re_wikipedia_wikidata_li",
            )

    if "REdefinitions_author_vol" in workbook.sheetnames:
        sheet = workbook["REdefinitions_author_vol"]
        headers = [normalize_space(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = row_dict(headers, row)
            update_re_enrichment(
                table,
                data.get("formal"),
                short_definition=data.get("definition"),
                volume=data.get("Volume"),
                page=data.get("page"),
                author=data.get("author"),
                match_source="REdefinitions_author_vol",
            )

    if "RE cited by headword not WD" in workbook.sheetnames:
        sheet = workbook["RE cited by headword not WD"]
        headers = [normalize_space(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = row_dict(headers, row)
            update_re_enrichment(
                table,
                data.get("REHeadword"),
                subject_item=data.get("WD_ID") or data.get("item"),
                subject_label=data.get("itemLabel"),
                match_source="RE cited by headword not WD",
            )

    workbook.close()
    return table


def enrich_re_mentions(parsed: ParsedToposText, re_lookup: dict[str, ReEnrichment]) -> ParsedToposText:
    if not re_lookup:
        return parsed
    mentions: list[Mention] = []
    for mention in parsed.mentions:
        if mention.authority_class != "re":
            mentions.append(mention)
            continue
        enrichment = None
        for key in re_lookup_keys(mention.tag_id):
            enrichment = re_lookup.get(key)
            if enrichment:
                break
        if enrichment is None:
            mentions.append(replace(mention, re_namespace_id=normalize_re_id(mention.tag_id)))
            continue
        mentions.append(
            replace(
                mention,
                re_namespace_id=enrichment.re_id,
                re_short_definition=enrichment.short_definition,
                re_article_item=enrichment.article_item,
                re_subject_item=enrichment.subject_item,
                re_subject_label=enrichment.subject_label,
                re_author=enrichment.author,
                re_volume=enrichment.volume,
                re_page=enrichment.page,
                re_match_source=enrichment.match_source,
            )
        )
    return ParsedToposText(entries=parsed.entries, mentions=mentions)


def top_entries_by_unresolved(parsed: ParsedToposText, limit: int) -> list[dict[str, object]]:
    entry_lookup = {entry.sequence: entry for entry in parsed.entries}
    unresolved_classes = {"zzz", "yy_placeholder", "jj_placeholder", "missing", "other"}
    counts: Counter[int] = Counter()
    zzz_counts: Counter[int] = Counter()
    placeholder_counts: Counter[int] = Counter()

    for mention in parsed.mentions:
        if mention.authority_class in unresolved_classes:
            counts[mention.entry_sequence] += 1
        if mention.authority_class == "zzz":
            zzz_counts[mention.entry_sequence] += 1
        if mention.authority_class in {"yy_placeholder", "jj_placeholder"}:
            placeholder_counts[mention.entry_sequence] += 1

    rows = []
    for entry_sequence, unresolved_count in counts.most_common(limit):
        entry = entry_lookup[entry_sequence]
        rows.append(
            {
                "entry_key": entry.entry_key,
                "title": entry.title,
                "unresolved": unresolved_count,
                "zzz": zzz_counts[entry_sequence],
                "yy_jj": placeholder_counts[entry_sequence],
                "snippet": entry.text[:260] + ("..." if len(entry.text) > 260 else ""),
            }
        )
    return rows


def top_surface_forms(mentions: Iterable[Mention], authority_class: str, limit: int) -> list[dict[str, object]]:
    counts: Counter[str] = Counter()
    examples: dict[str, Mention] = {}
    for mention in mentions:
        if mention.authority_class != authority_class:
            continue
        label = mention.mention_text or "(empty text)"
        counts[label] += 1
        examples.setdefault(label, mention)
    rows = []
    for label, count in counts.most_common(limit):
        mention = examples[label]
        rows.append(
            {
                "surface": label,
                "count": count,
                "entry_key": mention.entry_key,
                "context": mention.context,
            }
        )
    return rows


def build_re_namespace_rows(mentions: Iterable[Mention]) -> list[dict[str, object]]:
    grouped: dict[str, dict[str, object]] = {}
    original_ids: dict[str, set[str]] = defaultdict(set)
    counts: Counter[str] = Counter()
    for mention in mentions:
        if mention.authority_class != "re":
            continue
        re_id = mention.re_namespace_id or normalize_re_id(mention.tag_id)
        counts[re_id] += 1
        original_ids[re_id].add(mention.tag_id)
        grouped.setdefault(
            re_id,
            {
                "re_namespace_id": re_id,
                "count": 0,
                "original_ids": "",
                "short_definition": mention.re_short_definition,
                "re_article_item": mention.re_article_item,
                "subject_item": mention.re_subject_item,
                "subject_label": mention.re_subject_label,
                "author": mention.re_author,
                "volume": mention.re_volume,
                "page": mention.re_page,
                "match_source": mention.re_match_source,
                "first_entry": mention.entry_key,
                "first_text": mention.mention_text,
                "first_context": mention.context,
            },
        )
        row = grouped[re_id]
        for key, value in {
            "short_definition": mention.re_short_definition,
            "re_article_item": mention.re_article_item,
            "subject_item": mention.re_subject_item,
            "subject_label": mention.re_subject_label,
            "author": mention.re_author,
            "volume": mention.re_volume,
            "page": mention.re_page,
            "match_source": mention.re_match_source,
        }.items():
            if value and not row.get(key):
                row[key] = value

    rows = []
    for re_id, row in grouped.items():
        row["count"] = counts[re_id]
        row["original_ids"] = "; ".join(sorted(original_ids[re_id]))
        rows.append(row)
    return sorted(rows, key=lambda row: (-int(row["count"]), str(row["re_namespace_id"])))


def example_mentions(
    mentions: Iterable[Mention],
    authority_classes: set[str] | None = None,
    tag_names: set[str] | None = None,
    limit: int = 40,
) -> list[Mention]:
    rows = []
    for mention in mentions:
        if authority_classes is not None and mention.authority_class not in authority_classes:
            continue
        if tag_names is not None and mention.tag_name not in tag_names:
            continue
        rows.append(mention)
        if len(rows) >= limit:
            break
    return rows


def render_cell(value: object) -> str:
    return html_lib.escape("" if value is None else str(value))


def link_or_text(label: str, url: str) -> str:
    safe_label = render_cell(label)
    if not url:
        return safe_label
    return f"<a href=\"{render_cell(url)}\" target=\"_blank\" rel=\"noopener\">{safe_label}</a>"


def render_count_table(headers: list[str], rows: list[list[object]]) -> str:
    header_html = "".join(f"<th>{render_cell(header)}</th>" for header in headers)
    row_html = []
    for row in rows:
        row_html.append("<tr>" + "".join(f"<td>{render_cell(cell)}</td>" for cell in row) + "</tr>")
    return f"<table><thead><tr>{header_html}</tr></thead><tbody>{''.join(row_html)}</tbody></table>"


def render_mentions_table(rows: list[Mention]) -> str:
    if not rows:
        return "<p class=\"empty\">No examples in this snapshot.</p>"
    body = []
    for mention in rows:
        body.append(
            "<tr>"
            f"<td>{render_cell(mention.entry_key)}</td>"
            f"<td>{render_cell(mention.entry_title)}</td>"
            f"<td><code>{render_cell(mention.tag_name)}</code></td>"
            f"<td>{link_or_text(mention.tag_id, mention.authority_url)}</td>"
            f"<td>{render_cell(AUTHORITY_CLASS_LABELS.get(mention.authority_class, mention.authority_class))}</td>"
            f"<td>{render_cell(mention.mention_text)}</td>"
            f"<td>{render_cell(mention.context)}</td>"
            "</tr>"
        )
    return (
        "<table><thead><tr>"
        "<th>Entry</th><th>Title</th><th>Tag</th><th>ID</th><th>Class</th><th>Text</th><th>Context</th>"
        "</tr></thead><tbody>"
        + "".join(body)
        + "</tbody></table>"
    )


def render_dict_table(headers: list[str], rows: list[dict[str, object]], keys: list[str]) -> str:
    if not rows:
        return "<p class=\"empty\">No rows.</p>"
    header_html = "".join(f"<th>{render_cell(header)}</th>" for header in headers)
    body = []
    for row in rows:
        body.append("<tr>" + "".join(f"<td>{render_cell(row.get(key, ''))}</td>" for key in keys) + "</tr>")
    return f"<table><thead><tr>{header_html}</tr></thead><tbody>{''.join(body)}</tbody></table>"


def build_report_html(
    *,
    metadata: SnapshotMetadata,
    snapshot_path: Path,
    parsed: ParsedToposText,
    output_path: Path,
    mentions_csv_path: Path | None,
    re_namespace_csv_path: Path | None,
    pauly_workbook_path: Path | None,
    example_limit: int,
    top_limit: int,
) -> str:
    tag_counts = Counter(mention.tag_name for mention in parsed.mentions)
    class_counts = Counter(mention.authority_class for mention in parsed.mentions)
    work_counts = Counter(entry.work or "(missing)" for entry in parsed.entries)
    distinct_ids = len({mention.tag_id for mention in parsed.mentions if mention.tag_id})
    unknown_tag_counts = Counter(
        mention.tag_name for mention in parsed.mentions if mention.tag_name not in KNOWN_ENTITY_TAGS
    )
    tag_fix_counts = Counter(
        (mention.original_tag_name, mention.tag_name)
        for mention in parsed.mentions
        if mention.original_tag_name and mention.original_tag_name != mention.tag_name
    )
    by_tag_class = count_by_tag_and_class(parsed.mentions)
    re_namespace_rows = build_re_namespace_rows(parsed.mentions)
    re_distinct_count = len(re_namespace_rows)
    re_definition_count = sum(1 for row in re_namespace_rows if row.get("short_definition"))
    re_subject_count = sum(1 for row in re_namespace_rows if row.get("subject_item"))

    summary_cards = [
        ("Paragraphs", f"{len(parsed.entries):,}"),
        ("Entity Mentions", f"{len(parsed.mentions):,}"),
        ("Distinct IDs", f"{distinct_ids:,}"),
        ("zzz unresolved", f"{class_counts['zzz']:,}"),
        ("YY placeholders", f"{class_counts['yy_placeholder']:,}"),
        ("JJ placeholders", f"{class_counts['jj_placeholder']:,}"),
        ("RE references", f"{class_counts['re']:,}"),
        ("RE IDs matched", f"{re_definition_count:,}/{re_distinct_count:,}"),
        ("ToposText-like IDs", f"{class_counts['topostext_like']:,}"),
    ]

    tag_rows = []
    for tag_name, total in tag_counts.most_common():
        class_counter = by_tag_class[tag_name]
        tag_ids = {
            mention.tag_id
            for mention in parsed.mentions
            if mention.tag_name == tag_name and mention.tag_id
        }
        tag_rows.append(
            [
                tag_name,
                total,
                len(tag_ids),
                class_counter["zzz"],
                class_counter["yy_placeholder"],
                class_counter["jj_placeholder"],
                class_counter["re"],
                class_counter["wikidata"],
                class_counter["pleiades_numeric"],
                class_counter["topostext_like"],
            ]
        )

    class_rows = [
        [
            AUTHORITY_CLASS_LABELS.get(authority_class, authority_class),
            count,
            AUTHORITY_CLASS_NOTES.get(authority_class, ""),
        ]
        for authority_class, count in class_counts.most_common()
    ]

    work_rows = [[work, count] for work, count in work_counts.most_common()]
    unknown_tag_rows = [[tag, count] for tag, count in unknown_tag_counts.most_common()]
    tag_fix_rows = [[source, target, count] for (source, target), count in sorted(tag_fix_counts.items())]

    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    css = """
    :root {
      color-scheme: light;
      --ink: #1f2933;
      --muted: #586575;
      --line: #d7dde5;
      --panel: #f7f9fb;
      --accent: #165a72;
      --warn: #8a4b08;
    }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      color: var(--ink);
      background: #fff;
      line-height: 1.45;
    }
    main {
      max-width: 1220px;
      margin: 0 auto;
      padding: 32px 28px 56px;
    }
    h1, h2, h3 { line-height: 1.2; margin: 0; }
    h1 { font-size: 30px; margin-bottom: 8px; }
    h2 { font-size: 22px; margin-top: 36px; margin-bottom: 12px; border-bottom: 1px solid var(--line); padding-bottom: 8px; }
    h3 { font-size: 17px; margin-top: 24px; margin-bottom: 8px; }
    p { margin: 8px 0 14px; }
    code { font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; font-size: 0.92em; }
    a { color: var(--accent); text-decoration-thickness: 1px; text-underline-offset: 2px; }
    .metadata {
      color: var(--muted);
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
      gap: 4px 22px;
      margin: 14px 0 24px;
      font-size: 14px;
    }
    .cards {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
      gap: 10px;
      margin: 22px 0 28px;
    }
    .card {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--panel);
      padding: 12px 14px;
    }
    .card strong {
      display: block;
      font-size: 24px;
      margin-bottom: 2px;
    }
    .card span { color: var(--muted); font-size: 13px; }
    .note {
      border-left: 4px solid var(--accent);
      background: #eef6f8;
      padding: 12px 14px;
      margin: 16px 0 20px;
    }
    .warning {
      border-left-color: var(--warn);
      background: #fff7eb;
    }
    table {
      border-collapse: collapse;
      width: 100%;
      margin: 10px 0 20px;
      font-size: 13px;
    }
    th, td {
      border: 1px solid var(--line);
      padding: 7px 8px;
      vertical-align: top;
      text-align: left;
    }
    th {
      background: #eef2f6;
      font-weight: 650;
    }
    td:nth-child(1), td:nth-child(3), td:nth-child(4), td:nth-child(5) {
      white-space: nowrap;
    }
    .empty { color: var(--muted); font-style: italic; }
    ul { margin: 8px 0 16px 22px; padding: 0; }
    li { margin: 5px 0; }
    """

    cards_html = "".join(
        f"<div class=\"card\"><strong>{render_cell(value)}</strong><span>{render_cell(label)}</span></div>"
        for label, value in summary_cards
    )

    csv_html = ""
    if mentions_csv_path:
        relative_csv = mentions_csv_path.name if mentions_csv_path.parent == output_path.parent else str(mentions_csv_path)
        csv_html = (
            f"<p>Complete mention CSV: <code>{render_cell(str(mentions_csv_path))}</code> "
            f"(<a href=\"{render_cell(relative_csv)}\">open CSV next to this report</a>).</p>"
        )
    if re_namespace_csv_path:
        relative_re_csv = (
            re_namespace_csv_path.name
            if re_namespace_csv_path.parent == output_path.parent
            else str(re_namespace_csv_path)
        )
        csv_html += (
            f"<p>RE namespace CSV: <code>{render_cell(str(re_namespace_csv_path))}</code> "
            f"(<a href=\"{render_cell(relative_re_csv)}\">open RE namespace CSV next to this report</a>).</p>"
        )

    pauly_note = (
        f"PaulyHeadwords enrichment workbook: <code>{render_cell(pauly_workbook_path.name)}</code>."
        if pauly_workbook_path
        else "No PaulyHeadwords workbook was loaded; RE namespace rows are still listed, but German definitions and subject Wikidata enrichment are blank."
    )

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>ToposText Stephanus Intake Report</title>
  <style>{css}</style>
</head>
<body>
<main>
  <h1>ToposText Stephanus Intake Report</h1>
  <p>This report parses Brady's current ToposText HTML snapshot without writing database rows. It is meant to expose the source shape and unresolved entity decisions before we design the durable import tables.</p>

  <div class="metadata">
    <div><strong>Generated:</strong> {render_cell(generated_at)} UTC</div>
    <div><strong>Snapshot ID:</strong> {render_cell(metadata.snapshot_id or "local file")}</div>
    <div><strong>Snapshot status:</strong> {render_cell(metadata.status)}</div>
    <div><strong>Source name:</strong> {render_cell(metadata.source_name)}</div>
    <div><strong>Source kind:</strong> {render_cell(metadata.source_kind)}</div>
    <div><strong>Snapshot path:</strong> <code>{render_cell(str(snapshot_path))}</code></div>
    <div><strong>SHA-256:</strong> <code>{render_cell(metadata.sha256)}</code></div>
    <div><strong>Bytes:</strong> {render_cell(f"{metadata.byte_count:,}")}</div>
  </div>

  <div class="cards">{cards_html}</div>
  {csv_html}

  <section class="note">
    <h2>Current Working Assumptions</h2>
    <ul>
      <li><code>work=241</code> and paragraph <code>id</code> are treated as stable source coordinates for this import slice.</li>
      <li><code>Q...</code> is Wikidata; bare numeric IDs are treated as Pleiades-like candidates; <code>RE:...</code> is its own authority namespace.</li>
      <li>Mixed numeric/alphabetic IDs such as <code>386229PAba</code> are treated as ToposText-like keys and linked using the repo's existing <code>https://topostext.org/place/{{id}}</code> convention.</li>
      <li><code>YY</code> means Brady did a quick pattern search for a ToposText ID, but did not manually search Wikidata, RE, or Pleiades.</li>
      <li><code>JJ</code> means Brady searched and found nothing; we need to mint a fresh ToposText-style ID from the best guess of what and where it is.</li>
      <li><code>zzz</code> is reported as unresolved inline markup, not as a real authority identifier.</li>
      <li>Source tag typos are normalized on intake: <code>PPN</code> becomes <code>PRN</code>, and <code>demonymn</code> becomes <code>demonym</code>.</li>
    </ul>
  </section>

  <h2>Counts By Work</h2>
  {render_count_table(["work", "paragraphs"], work_rows)}

  <h2>Counts By Tag</h2>
  {render_count_table(["tag", "mentions", "distinct IDs", "zzz", "YY", "JJ", "RE", "Wikidata", "numeric", "ToposText-like"], tag_rows)}

  <h2>Authority Classes</h2>
  {render_count_table(["class", "mentions", "current interpretation"], class_rows)}

  <h2>Tag Variants To Check</h2>
  <h3>Source Tag Typos Normalized</h3>
  {render_count_table(["source tag", "canonical tag", "mentions"], tag_fix_rows) if tag_fix_rows else '<p class="empty">No known source tag typos found.</p>'}
  <h3>Unknown ID-Bearing Tags</h3>
  <p>Anything here has an <code>id</code> attribute but is not one of the expected canonical entity tags: {render_cell(", ".join(sorted(KNOWN_ENTITY_TAGS)))}.</p>
  {render_count_table(["tag", "mentions"], unknown_tag_rows) if unknown_tag_rows else '<p class="empty">No unknown id-bearing entity tag names found.</p>'}

  <h2>RE Namespace Enrichment</h2>
  <p>{pauly_note}</p>
  {render_count_table(["metric", "count"], [["RE mentions", class_counts["re"]], ["distinct RE namespace IDs", re_distinct_count], ["IDs with German short definition", re_definition_count], ["IDs with subject Wikidata item", re_subject_count], ["IDs without matched definition", re_distinct_count - re_definition_count]])}
  {render_dict_table(["RE namespace ID", "mentions", "original IDs", "German short definition", "subject Wikidata item", "article item", "first entry"], re_namespace_rows[:top_limit], ["re_namespace_id", "count", "original_ids", "short_definition", "subject_item", "re_article_item", "first_entry"])}

  <h2>Review Queues</h2>
  <h3>YY / JJ Placeholder Examples</h3>
  {render_mentions_table(example_mentions(parsed.mentions, {"yy_placeholder", "jj_placeholder"}, limit=example_limit))}

  <h3>RE Reference Examples</h3>
  {render_mentions_table(example_mentions(parsed.mentions, {"re"}, limit=example_limit))}

  <h3>Numeric / Pleiades-Like Examples</h3>
  {render_mentions_table(example_mentions(parsed.mentions, {"pleiades_numeric"}, limit=example_limit))}

  <h3>Other ID Examples</h3>
  {render_mentions_table(example_mentions(parsed.mentions, {"other"}, limit=example_limit))}

  <h3>Unknown Tag Name Examples</h3>
  {render_mentions_table(example_mentions(parsed.mentions, tag_names=set(unknown_tag_counts), limit=example_limit))}

  <h2>Top Unresolved <code>zzz</code> Surface Forms</h2>
  {render_dict_table(["surface", "count", "first entry", "first context"], top_surface_forms(parsed.mentions, "zzz", top_limit), ["surface", "count", "entry_key", "context"])}

  <h2>Entries With Most Unresolved Markup</h2>
  {render_dict_table(["entry", "title", "unresolved", "zzz", "YY/JJ", "snippet"], top_entries_by_unresolved(parsed, top_limit), ["entry_key", "title", "unresolved", "zzz", "yy_jj", "snippet"])}

  <section class="note warning">
    <h2>What This Does Not Yet Do</h2>
    <p>It does not mutate Stephanos entity-resolution tables, import rows into PostgreSQL, or decide canonical entity clusters. The next implementation step should use this report to settle the staging schema and review workflow before writing import data back to the database.</p>
  </section>
</main>
</body>
</html>
"""


def write_mentions_csv(path: Path, mentions: list[Mention]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "sequence",
        "work",
        "paragraph_id",
        "entry_key",
        "entry_title",
        "wdate",
        "edate",
        "tag_name",
        "original_tag_name",
        "tag_id",
        "authority_class",
        "mention_text",
        "authority_url",
        "context",
        "re_namespace_id",
        "re_short_definition",
        "re_article_item",
        "re_subject_item",
        "re_subject_label",
        "re_author",
        "re_volume",
        "re_page",
        "re_match_source",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for mention in mentions:
            row = {name: getattr(mention, name) for name in fieldnames}
            writer.writerow(row)


def write_re_namespace_csv(path: Path, mentions: list[Mention]) -> None:
    rows = build_re_namespace_rows(mentions)
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "re_namespace_id",
        "count",
        "original_ids",
        "short_definition",
        "re_article_item",
        "subject_item",
        "subject_label",
        "author",
        "volume",
        "page",
        "match_source",
        "first_entry",
        "first_text",
        "first_context",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_summary_json(path: Path, metadata: SnapshotMetadata, parsed: ParsedToposText) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    re_namespace_rows = build_re_namespace_rows(parsed.mentions)
    payload = {
        "metadata": asdict(metadata),
        "counts": {
            "entries": len(parsed.entries),
            "mentions": len(parsed.mentions),
            "by_tag": Counter(mention.tag_name for mention in parsed.mentions),
            "by_authority_class": Counter(mention.authority_class for mention in parsed.mentions),
            "by_work": Counter(entry.work or "(missing)" for entry in parsed.entries),
            "re_namespace_ids": len(re_namespace_rows),
            "re_namespace_ids_with_short_definition": sum(1 for row in re_namespace_rows if row.get("short_definition")),
            "re_namespace_ids_with_subject_item": sum(1 for row in re_namespace_rows if row.get("subject_item")),
            "source_tag_fixes": Counter(
                f"{mention.original_tag_name}->{mention.tag_name}"
                for mention in parsed.mentions
                if mention.original_tag_name and mention.original_tag_name != mention.tag_name
            ),
        },
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a ToposText HTML intake report.")
    parser.add_argument("--snapshot-file", type=Path, help="Parse this local HTML file instead of querying the DB")
    parser.add_argument("--snapshot-id", type=int, help="Specific entity_source_snapshots id to parse")
    parser.add_argument("--source-name", default=DEFAULT_SOURCE_NAME)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--mentions-csv",
        type=Path,
        help="Complete mention CSV path. Defaults to <output stem>_mentions.csv",
    )
    parser.add_argument(
        "--re-namespace-csv",
        type=Path,
        help="Distinct RE namespace CSV path. Defaults to <output stem>_re_namespace.csv",
    )
    parser.add_argument("--no-mentions-csv", action="store_true", help="Do not write the complete mention CSV")
    parser.add_argument("--no-re-namespace-csv", action="store_true", help="Do not write the RE namespace CSV")
    parser.add_argument("--summary-json", type=Path, help="Optional machine-readable summary JSON path")
    parser.add_argument("--pauly-workbook", type=Path, help="PaulyHeadwords workbook for RE enrichment")
    parser.add_argument("--no-pauly-enrichment", action="store_true", help="Skip PaulyHeadwords workbook lookup")
    parser.add_argument("--example-limit", type=int, default=40)
    parser.add_argument("--top-limit", type=int, default=30)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    args.output = args.output.expanduser()

    if args.snapshot_file:
        snapshot_path = args.snapshot_file.expanduser()
        metadata = metadata_from_file(snapshot_path, args.source_name)
    else:
        metadata = latest_snapshot_from_db(args.source_name, snapshot_id=args.snapshot_id)
        snapshot_path = resolve_snapshot_path(metadata)

    content = snapshot_path.read_text(encoding="utf-8")
    parsed = parse_topostext_html(content)

    pauly_workbook_path = None
    if not args.no_pauly_enrichment:
        if args.pauly_workbook:
            pauly_workbook_path = args.pauly_workbook.expanduser()
            if not pauly_workbook_path.exists():
                raise FileNotFoundError(f"Pauly workbook not found: {pauly_workbook_path}")
        else:
            pauly_workbook_path = find_default_pauly_workbook()
        if pauly_workbook_path:
            parsed = enrich_re_mentions(parsed, load_pauly_re_enrichment(pauly_workbook_path))

    mentions_csv_path = None
    if not args.no_mentions_csv:
        mentions_csv_path = (
            args.mentions_csv.expanduser()
            if args.mentions_csv
            else args.output.with_name(f"{args.output.stem}_mentions.csv")
        )
        write_mentions_csv(mentions_csv_path, parsed.mentions)

    re_namespace_csv_path = None
    if not args.no_re_namespace_csv:
        re_namespace_csv_path = (
            args.re_namespace_csv.expanduser()
            if args.re_namespace_csv
            else args.output.with_name(f"{args.output.stem}_re_namespace.csv")
        )
        write_re_namespace_csv(re_namespace_csv_path, parsed.mentions)

    report_html = build_report_html(
        metadata=metadata,
        snapshot_path=snapshot_path,
        parsed=parsed,
        output_path=args.output,
        mentions_csv_path=mentions_csv_path,
        re_namespace_csv_path=re_namespace_csv_path,
        pauly_workbook_path=pauly_workbook_path,
        example_limit=args.example_limit,
        top_limit=args.top_limit,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(report_html, encoding="utf-8")

    if args.summary_json:
        write_summary_json(args.summary_json.expanduser(), metadata, parsed)

    print(f"entries={len(parsed.entries)}")
    print(f"mentions={len(parsed.mentions)}")
    print(f"report={args.output}")
    if mentions_csv_path:
        print(f"mentions_csv={mentions_csv_path}")
    if re_namespace_csv_path:
        print(f"re_namespace_csv={re_namespace_csv_path}")
    if pauly_workbook_path:
        print(f"pauly_workbook={pauly_workbook_path}")
    if args.summary_json:
        print(f"summary_json={args.summary_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
